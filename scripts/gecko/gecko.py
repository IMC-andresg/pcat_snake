from __future__ import print_function, unicode_literals
import json
import logging
import os
import time
from pathlib import Path
import pandas as pd
import requests
import tabulate
from PyInquirer import prompt, print_json
from openpyxl import load_workbook

from shutil import copy2, rmtree

MAX_ITEM_NAME = 60

Name_Shorts = {
    "Microsoft 365": "M365",
    "Office 365": "O365",
    "Services": "Svcs",
    "Device": "Dvc",
    "Windows": "W"
}
# Read config

with open('config.json') as json_data_file:
	config = json.load(json_data_file)

# Ask User for Target Product

questions = [
    {
        'type': 'input',
        'name': 'project_name',
        'message': 'Project name',
    },
    {
        'type': 'list',
        'name': 'country',
        'message': 'Country',
        'choices':['US','CA','GR','MX','BR','LA','CO','NL','UK','FR','DE','IT','ES','SE','BE','AT','CH','NO','ME','TR','MA','AU','NZ','SG','HK','IN','MY','ZA', 'PL']
    },
    {
        'type': 'list',
        'name': 'product_name',
        'message': 'product',
        'choices': config['CONNECT_PRODUCTS'].keys()
    }
]
answers = prompt(questions)
config_product = config['CONNECT_PRODUCTS'][answers['product_name']]
product_id = config_product['ID']
project_name = answers['project_name']
output_path = "./out/" + project_name
country = answers['country']

# Load Connect Offers

def billing_period_from_name(name):
    if config_product['Plan_Billing_Period'] < 0:
        if '1 year' in name:
            return 1
        elif '3 year' in name:
            return 3
        else:
            return 1
    else:
        return config_product['Plan_Billing_Period']

def shorten_name(name): 
    new_name = name
    if len(new_name) >= MAX_ITEM_NAME:
        for i, j in Name_Shorts.items():
            new_name = new_name.replace(i, j)
            if len(new_name) < MAX_ITEM_NAME:
                break
    if len(new_name) >= MAX_ITEM_NAME:
        raise Exception("Could not shorten item name further "+ new_name)
    
    return new_name

def load_connect_items():
    connect_items = pd.DataFrame()
    headers = {'Authorization': config['CONNECT_TOKEN']}
    # url = config['CONNECT_URLS']['PRODUCT'].replace('<PRODUCT_ID>', product_id)
    # connect_product = requests.get(url, headers=headers).json()
    url = config['CONNECT_URLS']['ITEMS'].replace('<PRODUCT_ID>', product_id)
    offset = 0
    params = {'limit': config['CONNECT_PARAMS']['LIMIT'], 'offset': offset}
    response = requests.get(url, headers=headers, params=params).json()

    while response:
        for item_in_product in response:
            sku = item_in_product['mpn']
            connect_items.loc[sku, 'mpn'] = sku
            # connect_items.loc[sku, 'name'] = shorten_name(item_in_product['name'])
            connect_items.loc[sku, 'name'] = item_in_product['name']
            connect_items.loc[sku, 'description'] = item_in_product['description']
            connect_items.loc[sku, 'billing_period'] = billing_period_from_name(item_in_product['name'])

# TODO add special max 500 quantity for DG7GMGF0DVSV:000P:SoftwareSubscriptions, DG7GMGF0DVSV:000G:SoftwareSubscriptions, DG7GMGF0DVSV:000L:SoftwarePerpetual, DG7GMGF0DVSV:000F:SoftwarePerpetual

        offset = offset + config['CONNECT_PARAMS']['LIMIT']
        params = {'limit': config['CONNECT_PARAMS']['LIMIT'], 'offset': offset}
        response = requests.get(url, headers=headers, params=params).json()
    connect_items.sort_values(by=['name'], ascending=True, inplace=True)
    return connect_items

connect_items = load_connect_items()
total = len(connect_items) 

# Load Excel Templates
l0_path = output_path+"/"+config_product['L0_TEMPLATE'].replace('template',project_name)
l1_path = output_path+"/"+config_product['L1_TEMPLATE'].replace('template',project_name)

# if os.path.exists(project_name):
#     rmtree(project_name)

if not os.path.exists("./out/"):
    os.mkdir("./out/")

if not os.path.exists(output_path):
    os.mkdir(output_path)

copy2(config['TEMPLATES_PATH']+config_product['L0_TEMPLATE'], l0_path)
copy2(config['TEMPLATES_PATH']+config_product['L1_TEMPLATE'], l1_path)

# Generate Date & Save

def fill_col(writer, sheet, value, row, col):
    pd.concat([pd.DataFrame([value])]*total).to_excel(writer, sheet, index=False,header=False, startrow=row, startcol=col)

## L0

with pd.ExcelWriter(l0_path, engine='openpyxl', mode='a') as writer: # pylint: disable=abstract-class-instantiated
    writer.book = load_workbook(l0_path)
    writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
    # Plan Sheet
    connect_items['name'].to_excel(writer, "Plan", index=False,header=False, startrow=1, startcol=1)
    connect_items['description'].to_excel(writer, "Plan", index=False,header=False, startrow=1, startcol=6)
    fill_col(writer, "Plan", config_product['Plan_Template'], 1, 2)
    fill_col(writer, "Plan", config_product['Plan_Plan_Category'], 1, 3)
    fill_col(writer, "Plan", config_product['Plan_Service_Term'], 1, 5)
    fill_col(writer, "Plan", config_product['Plan_Notification_Tmpl'], 1, 7)
    fill_col(writer, "Plan", config_product['Plan_Billing_Period_Type'], 1, 9)
    connect_items['billing_period'].to_excel(writer, "Plan", index=False,header=False, startrow=1, startcol=10)
    fill_col(writer, "Plan", config_product['Plan_Price_Period'], 1, 11)
    fill_col(writer, "Plan", config_product['Plan_Recurring_Type'], 1, 12)
    fill_col(writer, "Plan", config_product['Plan_Auto_Renew'], 1, 13)
    fill_col(writer, "Plan", config_product['Plan_Renew_Order_Interval'], 1, 14)
    fill_col(writer, "Plan", config_product['Plan_Sales_Category'], 1, 16)

    #Period
    connect_items['name'].to_excel(writer, "Period", index=False,header=False, startrow=1, startcol=1)
    connect_items['billing_period'].to_excel(writer, "Period", index=False,header=False, startrow=1, startcol=2)

    fill_col(writer, "Period", config_product['Plan_Billing_Period_Type'], 1, 3)
    fill_col(writer, "Period", config_product['Period_Cancellation_Type'], 1, 4)
    fill_col(writer, "Period", config_product['Period_Refund_Period'], 1, 5)
    fill_col(writer, "Period", config_product['Period_After_Refund'], 1, 6)
    fill_col(writer, "Period", config_product['Period_Trial'], 1, 8)
    
    #Resources
    connect_items['name'].to_excel(writer, "Resources", index=False,header=False, startrow=1, startcol=0)
    fill_col(writer, "Resources", config_product['Resources_ResCategory'], 1, 1)
    connect_items['mpn'].to_excel(writer, "Resources", index=False,header=False, startrow=1, startcol=3)
    fill_col(writer, "Resources", config_product['Resources_UOM'], 1, 4)


    #Resource Rates
    connect_items['name'].to_excel(writer, "Resource Rates", index=False,header=False, startrow=1, startcol=1)
    connect_items['name'].to_excel(writer, "Resource Rates", index=False,header=False, startrow=1, startcol=2)
    fill_col(writer, "Resource Rates", config_product['RRates_IncAmount'], 1, 4)
    fill_col(writer, "Resource Rates", config_product['RRates_MinUnits'], 1, 5)
    fill_col(writer, "Resource Rates", config_product['RRates_MaxAmount'], 1, 6)
    fill_col(writer, "Resource Rates", config_product['RRates_ShowInStore'], 1, 7)
    fill_col(writer, "Resource Rates", config_product['RRates_Measurable'], 1, 8)
    fill_col(writer, "Resource Rates", config_product['RRates_ShowInCP'], 1, 9)


## L1

with pd.ExcelWriter(l1_path, engine='openpyxl', mode='a') as writer: # pylint: disable=abstract-class-instantiated
    writer.book = load_workbook(l1_path)
    writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
    # Plan Sheet
    connect_items['name'].to_excel(writer, "Plan", index=False,header=False, startrow=1, startcol=0)
    fill_col(writer, "Plan", config_product['TermCondition'], 1, 2)
    fill_col(writer, "Plan", config_product['TermCondition2'], 1, 3)
    fill_col(writer, "Plan", config_product['TermCondition3'], 1, 4)

    #Resources
    connect_items['name'].to_excel(writer, "Resources", index=False,header=False, startrow=1, startcol=0)
    connect_items['mpn'].to_excel(writer, "Resources", index=False,header=False, startrow=1, startcol=1)

    #Resource Rates
    connect_items['name'].to_excel(writer, "Resource Rates", index=False,header=False, startrow=1, startcol=1)
    connect_items['name'].to_excel(writer, "Resource Rates", index=False,header=False, startrow=1, startcol=2)
    fill_col(writer, "Resource Rates", country, 1, 3)
    fill_col(writer, "Resource Rates", config_product['ModInTrial'], 1, 6)

# Done 
